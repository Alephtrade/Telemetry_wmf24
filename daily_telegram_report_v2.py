import json
import requests
import time
import logging
import telegram.strings as tg_strings
from datetime import timedelta, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from telegram.models import WMFTelegramBot
from db.models import WMFSQLDriver
from core.utils import timedelta_str, print_exception, get_env_mode, initialize_logger, \
    get_next_date_formed, get_curr_time_str, get_curr_time
from wmf.models import WMFMachineStatConnector
if get_env_mode() == 'prod':
    from settings import prod as settings
else:
    from settings import test as settings


time.sleep(5)
initialize_logger('daily_telegram_report_v2.log')
logging.info('Started daily_telegram_report_v2.py')
tg_conn = WMFTelegramBot()
db_conn = WMFSQLDriver()

wb = Workbook()
ws = wb.active
CW_MULTIPLIER = 1.5
redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

yellowFill = PatternFill(start_color='FFFF00',
                   end_color='FFFF00',
                   fill_type='solid')

boldFont = Font(bold=True)

t = ws.cell(1, 1, 'Ошибка')
t.fill = yellowFill
t.font = boldFont

t = ws.cell(1, 2, 'Наименование')
t.fill = yellowFill
t.font = boldFont

t = ws.cell(1, 3, 'Начало')
t.fill = yellowFill
t.font = boldFont

t = ws.cell(1, 4, 'Конец')
t.fill = yellowFill
t.font = boldFont

t = ws.cell(1, 5, 'Длительность (секунд)')
t.fill = yellowFill
t.font = boldFont

error_text_max_len = 12
ws.column_dimensions['A'].width = 6 * CW_MULTIPLIER
ws.column_dimensions['C'].width = 17 * CW_MULTIPLIER
ws.column_dimensions['D'].width = 17 * CW_MULTIPLIER
ws.column_dimensions['E'].width = 21 * CW_MULTIPLIER

try:
    def create_bolvanka():
        tg_last_report = db_conn.get_last_tg_report()
        if tg_last_report:
            last_date_formed = datetime.strptime(tg_last_report[1], '%Y-%m-%d %H:%M:%S')
            if get_curr_time() < last_date_formed or not tg_last_report[2]:
                return False

        next_date_formed = get_next_date_formed(settings.TELEGRAM_REPORT_INTERVAL_MINUTES)
        next_date_formed_str = next_date_formed.strftime('%Y-%m-%d %H:%M:%S')
        logging.info(f'Creating bolvanka in tg_reports with date_formed: {next_date_formed_str}')
        db_conn.create_tg_report(date_formed=next_date_formed_str, body='')
        return True

    def generate_tg_report():
        global error_text_max_len
        last_reports = db_conn.get_last_two_tg_reports()
        logging.info(f'generate_tg_report(): last two reports are\n{last_reports}')
        if not last_reports:
            return False

        time_worked = timedelta(minutes=settings.TELEGRAM_REPORT_INTERVAL_MINUTES)
        date_formed = get_curr_time()
        date_formed_str = date_formed.strftime('%Y-%m-%d %H:%M:%S')

        if last_reports[0][2] or datetime.strptime(last_reports[0][1], '%Y-%m-%d %H:%M:%S') > date_formed:
            return False

        if len(last_reports) == 2:
            time_worked = date_formed - datetime.strptime(last_reports[1][1], '%Y-%m-%d %H:%M:%S')

        wm_conn = WMFMachineStatConnector()
        if not wm_conn.ws:
            return False
        logging.info(f'Successfully connected to machine, part number is {wm_conn.part_number}')
        data = wm_conn.get_wmf_machine_info()
        stoppage_time, wmf_error_time = timedelta(), timedelta()
        stoppage_count, wmf_error_count = 0, 0
        row_counter = 2
        unsent_records = db_conn.get_unsent_records()
        for rec_id, error_code, start_time, end_time, error_text in unsent_records:
            if end_time:
                error_text = error_text if error_text else ''
                duration_time = datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S') - datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
                time_worked -= duration_time
                if error_code == -1:
                    stoppage_count += 1
                    stoppage_time += duration_time
                else:
                    wmf_error_count += 1
                    wmf_error_time += duration_time
                db_conn.set_report_sent(rec_id)

                ws.cell(row_counter, 1, error_code)
                ws.cell(row_counter, 2, error_text)
                error_text_max_len = max(error_text_max_len, len(error_text))
                ws.cell(row_counter, 3, start_time)
                ws.cell(row_counter, 4, end_time)
                ws.cell(row_counter, 5, duration_time.total_seconds())
                row_counter += 1

        ws.column_dimensions['B'].width = error_text_max_len * CW_MULTIPLIER
        wb.save('error_report.xlsx')
        wb.close()

        data['date_formed'] = date_formed_str
        data['time_worked'] = timedelta_str(time_worked)
        data['wmf_error_time'] = timedelta_str(wmf_error_time)
        data['wmf_error_count'] = wmf_error_count
        data['stoppage_time'] = timedelta_str(stoppage_time)
        data['stoppage_count'] = stoppage_count

        last_record = db_conn.get_last_record()
        last_bev_count, last_cleaning_datetime = last_record[0], last_record[2]
        curr_bev_count = wm_conn.get_beverages_count()
        if curr_bev_count:
            data['beverages_count'] = curr_bev_count - last_bev_count
            db_conn.save_last_record('beverages_count', curr_bev_count)
        else:
            data['beverages_count'] = tg_strings.NO_MACHINE_CONNECTION
        data['last_cleaning_datetime'] = last_cleaning_datetime

        tg_report_body = tg_strings.DAILY_REPORT.format(**data)
        data['code'] = wm_conn.part_number
        with open('error_report.json', 'w') as f:
            json.dump(data, f)
        db_conn.set_tg_report_body(tg_id=last_reports[0][0], date_formed=date_formed_str, body=tg_report_body)
        return True


    def send_tg_report():
        last_tg_report = db_conn.get_last_tg_report()
        if not last_tg_report:
            return
        if last_tg_report[2]:
            return
        if tg_conn.send_message(last_tg_report[3]):
            try:
                if load_workbook(filename='error_report.xlsx').active.max_row > 1:
                    with open('machine_info.txt') as f:
                        tg_conn.send_file('error_report.xlsx', f.read())
                r = requests.post(
                    settings.WMF_BASE_URL + '/api/report',
                    data=json.load(open('error_report.json')),
                    files={'file': open('error_report.xlsx', 'rb')}
                )
                content = r.content.decode('utf-8')
                logging.info(f'daily_telegram_report_v2 send_report to wmf24: => {r} {content}')
            except Exception as ex:
                logging.error(f'daily_telegram_report_v2 send_report to wmf24: ERROR={ex}')
                logging.error(print_exception())
            db_conn.set_tg_report_sent(tg_id=last_tg_report[0], date_sent=get_curr_time_str())


    if not create_bolvanka():
        time.sleep(5)
    if not generate_tg_report():
        time.sleep(5)
    send_tg_report()

    db_conn.close()
except Exception:
    tg_conn.send_message(print_exception(), debug_mode=True)
